#!/usr/bin/env python3
"""Build BoatBuilder's browser catalog from the maintained Google workbook."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import urllib.request
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

SHEET_ID = "17-WMY8q2cCw7smmwLoMWzHbqZTlahi0Atsqe7gh7Wqs"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
DECADE_FIELDS = ["1980s Value", "1990s Value", "2000s Value", "2010s Value", "2020s Value"]
CATEGORIES = [
    {"id": "boats", "name": "Boats", "order": 10},
    {"id": "main-motors", "name": "Main Motors", "order": 20},
    {"id": "kickers", "name": "Kicker Motors", "order": 30},
    {"id": "bow-trolling-motors", "name": "Bow Trolling Motors", "order": 40},
    {"id": "downriggers", "name": "Downriggers", "order": 50},
    {"id": "electronics", "name": "Electronics & Navigation", "order": 60},
    {"id": "canvas", "name": "Bimini, Canvas & Covers", "order": 70},
    {"id": "electrical", "name": "Electrical Systems", "order": 80},
]
EQUIPMENT_CATEGORY = {
    "Main Motor": "main-motors",
    "Kicker": "kickers",
    "Bow Trolling Motor": "bow-trolling-motors",
    "Downrigger": "downriggers",
    "Electronics": "electronics",
    "Bimini / Canvas": "canvas",
    "Electrical": "electrical",
}
APPROVED_PHOTO_MATCHES = {
    "exact",
    "exact model and era",
    "exact model and close era",
    "exact model; close era",
    "exact model; current generation",
    "exact factory deck plan; later era",
}

EQUIPMENT_SUBTYPES = {
    "electronics": {
        "sonar": ("Fish Finders & Sonar", 10),
        "combo": ("Fish Finder / Chartplotter Combos", 20),
        "mfd": ("Chartplotters & Multifunction Displays", 30),
        "vhf": ("VHF Marine Radios", 40),
    },
    "electrical": {
        "batteries": ("Batteries & Banks", 10),
        "charging": ("Chargers & Charge Management", 20),
        "protection": ("Circuit Protection & Distribution", 30),
        "switching": ("Switching & Isolation", 40),
        "monitoring": ("Monitoring", 50),
        "wiring": ("Wiring & Installation", 60),
    },
}


def equipment_subtype(row: dict[str, Any], category_id: str) -> dict[str, Any] | None:
    name = " ".join(
        clean(row.get(field))
        for field in ("Display Name", "Manufacturer / System", "Model / Component")
    ).lower()
    role = clean(row.get("Specs / Role")).lower()

    subtype_id: str | None = None

    if category_id == "electronics":
        if "vhf" in name or "vhf" in role or "marine radio" in name:
            subtype_id = "vhf"
        elif any(token in name for token in ("apex", "solix", "xplore", "hds live", "hds pro", "axiom /", "axiom+")) \
                or "premium multifunction" in role or role.startswith("multifunction chartplotter"):
            subtype_id = "mfd"
        elif "no gps" in role or "no internal chartplotter" in role or "gps waypoint plotting" in role:
            subtype_id = "sonar"
        elif "chartplotter" in role or "gps plotter" in role or "fish finder / gps" in role or "fishfinder / gps" in role:
            subtype_id = "combo"
        else:
            subtype_id = "sonar"

    elif category_id == "electrical":
        if name.startswith("battery bank") or name.startswith("battery chemistry"):
            subtype_id = "batteries"
        elif any(token in name for token in ("smartshunt", "battery monitor", "monitoring")):
            subtype_id = "monitoring"
        elif any(token in name for token in ("add-a-battery", "battery switch", "low-voltage disconnect", "m-lvd")):
            subtype_id = "switching"
        elif any(token in name for token in ("circuit breaker", "breaker", "fuse block")):
            subtype_id = "protection"
        elif any(token in name for token in ("charger", "charging relay", "si-acr", "dc-dc", "charger inlet")):
            subtype_id = "charging"
        else:
            subtype_id = "wiring"

    if subtype_id is None:
        return None

    label, order = EQUIPMENT_SUBTYPES[category_id][subtype_id]
    return {"id": subtype_id, "name": label, "order": order}



def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def rows_from_sheet(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [clean(value) for value in next(iterator)]
    except StopIteration as exc:
        raise RuntimeError(f"{sheet_name} is empty") from exc

    rows: list[dict[str, Any]] = []
    for values in iterator:
        row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
        if any(clean(value) for value in row.values()):
            rows.append(row)
    return rows


def parse_money_tokens(value: Any) -> list[int]:
    text = clean(value).replace(",", "")
    if not text:
        return []
    values: list[int] = []
    for match in re.finditer(r"\$?\s*(\d+(?:\.\d+)?)\s*([kK]?)", text):
        token = match.group(0)
        suffix = match.group(2)
        if "$" not in token and not suffix:
            continue
        amount = float(match.group(1))
        if suffix:
            amount *= 1000
        values.append(round(amount))
    return values


def broad_range(row: dict[str, Any]) -> tuple[int | None, int | None]:
    values: list[int] = []
    for field in DECADE_FIELDS:
        values.extend(parse_money_tokens(row.get(field)))
    return (min(values), max(values)) if values else (None, None)


def finite_or_none(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def detail_rows(row: dict[str, Any], fields: Iterable[str]) -> list[dict[str, str]]:
    details: list[dict[str, str]] = []
    for field in fields:
        value = clean(row.get(field))
        if value and value != "—":
            details.append({"label": field, "value": value})
    return details


def build_catalog(source: Path) -> dict[str, Any]:
    workbook = load_workbook(source, data_only=True, read_only=True)
    boats = rows_from_sheet(workbook, "App Boats")
    equipment = rows_from_sheet(workbook, "App Equipment")
    photos = rows_from_sheet(workbook, "Boat Photos")

    photos_by_boat = {clean(row.get("Boat ID")): row for row in photos if clean(row.get("Boat ID"))}
    category_name = {entry["id"]: entry["name"] for entry in CATEGORIES}
    items: list[dict[str, Any]] = []

    boat_fields = [
        "Model Years / Era", "Recommendation", "Big-Water Suitability", "Layout", "Length", "Beam",
        "Chine / Bottom Width", "Dry Hull Weight", "Max / Bow Depth", "Cockpit / Interior Depth",
        "Deadrise", "Transom Height", "Transom Width", "Max HP", "Practical Working HP", "Persons",
        "Capacity Weight", "Fuel Capacity", "Bottom Thickness", "Side / Freeboard Thickness", "Construction",
        "Availability Under $14k", "Placement Reason", "Notes", "Interior Finish / Deck Material",
        "Interior Material Basis", "Washdown / Carpet Fit", *DECADE_FIELDS,
    ]

    for row in boats:
        boat_id = clean(row.get("Boat ID"))
        stable_key = clean(row.get("AppSheet Key"))
        if not boat_id or not stable_key:
            continue
        low, high = broad_range(row)
        photo_row = photos_by_boat.get(boat_id, {})
        photo_url = clean(photo_row.get("Representative Photo URL"))
        match_quality = clean(photo_row.get("Match Quality"))
        image = None
        if photo_url and match_quality.lower() in APPROVED_PHOTO_MATCHES:
            image = {
                "url": photo_url,
                "source": clean(photo_row.get("Photo Source Page")),
                "matchQuality": match_quality,
                "note": clean(photo_row.get("Curation Notes")),
            }
        model = (
            clean(row.get("Exact Model / Variant"))
            or clean(row.get("Variant / Size"))
            or clean(row.get("Model Family"))
            or clean(row.get("Display Name"))
        )
        items.append({
            "id": f"boat:{stable_key}",
            "categoryId": "boats",
            "categoryName": "Boats",
            "manufacturer": clean(row.get("Brand")) or "Unknown",
            "model": model,
            "displayName": clean(row.get("Display Name")) or boat_id,
            "subtitle": clean(row.get("Model Years / Era")),
            "badge": clean(row.get("Recommendation")),
            "lowPrice": low,
            "highPrice": high,
            "priceBasis": "Broad hull guidance across the listed model eras. Choose the applicable era when judging a specific listing.",
            "sourceUrl": clean(row.get("Source URL")),
            "image": image,
            "details": detail_rows(row, boat_fields),
        })

    equipment_fields = [
        "Era / Status", "Specs / Role", "Features / Controls", "Known Concerns", "Inspect Before Valuing",
        "Great Lakes / Fit", "Desirability", "Value Guidance", "Audit / Notes", *DECADE_FIELDS,
    ]
    for row in equipment:
        item_id = clean(row.get("Equipment ID"))
        category_id = EQUIPMENT_CATEGORY.get(clean(row.get("Category")))
        if not item_id or not category_id:
            continue
        subtype = equipment_subtype(row, category_id)
        broad_low, broad_high = broad_range(row)
        entered_low = finite_or_none(row.get("Est Low"))
        entered_high = finite_or_none(row.get("Est High"))
        low = entered_low if entered_low is not None else broad_low
        high = entered_high if entered_high is not None else broad_high
        items.append({
            "id": item_id,
            "categoryId": category_id,
            "categoryName": category_name[category_id],
            "subtypeId": subtype["id"] if subtype else None,
            "subtypeName": subtype["name"] if subtype else None,
            "subtypeOrder": subtype["order"] if subtype else None,
            "manufacturer": clean(row.get("Manufacturer / System")) or "Unknown",
            "model": clean(row.get("Model / Component")) or clean(row.get("Display Name")),
            "displayName": clean(row.get("Display Name")) or clean(row.get("Model / Component")),
            "subtitle": clean(row.get("Era / Status")),
            "badge": clean(row.get("Desirability")),
            "lowPrice": low,
            "highPrice": high,
            "priceBasis": (
                "Configured equipment estimate from the spreadsheet."
                if entered_low is not None or entered_high is not None
                else "Broad family and era guidance. Exact year, size, horsepower, and condition can narrow this range."
            ),
            "sourceUrl": clean(row.get("Source URL")),
            "image": None,
            "details": detail_rows(row, equipment_fields),
        })

    ids = [item["id"] for item in items]
    duplicates = sorted({item_id for item_id in ids if ids.count(item_id) > 1})
    if duplicates:
        raise RuntimeError(f"Duplicate item IDs: {duplicates[:10]}")
    if not items:
        raise RuntimeError("Catalog contains no items")

    boat_count = sum(item["categoryId"] == "boats" for item in items)
    equipment_count = len(items) - boat_count
    return {
        "schemaVersion": 2,
        "source": "Generated from Aluminum boat model review",
        "counts": {"items": len(items), "boats": boat_count, "equipment": equipment_count},
        "categories": CATEGORIES,
        "items": items,
    }


def download_workbook(destination: Path) -> None:
    request = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "BoatBuilder catalog builder"})
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = response.read()
    if not payload.startswith(b"PK"):
        raise RuntimeError("Google did not return an XLSX workbook")
    destination.write_bytes(payload)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, help="Use a local XLSX file instead of downloading Google Sheets")
    parser.add_argument("--output", type=Path, default=Path("data/catalog.js"))
    args = parser.parse_args()

    source = args.input or Path("/tmp/boatbuilder-source.xlsx")
    if args.input is None:
        download_workbook(source)

    catalog = build_catalog(source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    payload = "window.BOATBUILDER_DATA=" + json.dumps(catalog, ensure_ascii=False, separators=(",", ":")) + ";\n"
    args.output.write_text(payload, encoding="utf-8")
    print(json.dumps(catalog["counts"], sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
